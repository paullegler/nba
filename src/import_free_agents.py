from openpyxl import load_workbook

names = set()
path = "../files/players.txt"
with open(path) as f:
    content = f.readlines()
content = [x.strip() for x in content]
for name in content:
    names.add(name)

fname = "../excel/free_agents.xlsx"
output = "../files/free_agents.txt"
out = open(output, 'a')

wb = load_workbook(fname)
ws = wb['Sheet1']
for i in range(1, 350):
    for j in range(1, 30):
        name = ws.cell(i, j).value
        if name in names:
            out.write(name + '\n')

out.close()
